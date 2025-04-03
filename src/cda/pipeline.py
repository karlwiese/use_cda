from functools import lru_cache
from typing import Callable

import yaml

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.workbook import Workbook


SHEET_NAME_ATTRIBUTES: str = "Attributes"
SHEET_NAME_ENTITIES: str = "Entities"
SHEET_NAME_LICENSE: str = "License"

# own definitions
PICKLIST_ENTITIES_AND_ATTRIBUTES: dict[str, dict[str, str | list[dict]]] = {
    "Language Items": {
        "Name": "Language",
        "Label": "Language",
        "Description": "Language names follow the ISO 639 standard for language classification (https://en.wikipedia.org/wiki/List_of_ISO_639_language_codes). This list is a subset of the ISO 639 language codes based on frequency of usage in HCP data.",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "CHAR(2)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(20)",
                "Description": None,
            },
            {
                "Name": "direction",
                "Label": "Direction",
                "Data Type": "CHAR(3)",
                "Description": "Direction of reading",
            },
        ],
        "Values": [],
    },
    "Country Items": {
        "Name": "Country",
        "Label": "Country",
        "Description": "Based on ISO 3166-1 standard country codes and names (https://en.wikipedia.org/wiki/List_of_ISO_3166_country_codes#Current_ISO_3166_country_codes).",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "CHAR(2)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(80)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
        ],
        "Values": [],
    },
    "State Items": {
        "Name": "State",
        "Label": "State",
        "Description": "Based on ISO 3166-2 (https://en.wikipedia.org/wiki/ISO_3166-2) codes for identifying the principal subdivisions (e.g., provinces or states) of all countries coded in ISO 3166-1 (https://en.wikipedia.org/wiki/List_of_ISO_3166_country_codes#Current_ISO_3166_country_codes)",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "VARCHAR(6)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(80)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
        ],
        "Values": [],
    },
    "HCP Type Items": {
        "Name": "HCP_Type",
        "Label": "HCP Type",
        "Description": "The role an individual plays in the life sciences industry, spanning from the development and commercialization of life science products to their delivery and administration in healthcare settings.",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "CHAR(4)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(40)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
        ],
        "Values": [],
    },
    "Specialty Items": {
        "Name": "Specialty",
        "Label": "Specialty",
        "Description": "The primary medical field or expertise area to which the healthcare professional belongs. Uses the list of specialties.",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "CHAR(4)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(40)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
            {
                "Name": "specialty_group_mapping",
                "Label": "Specialty Group Mapping",
                "Data Type": "CHAR(2)",
                "Description": None,
            },
        ],
        "Values": [],
    },
    "Specialty Group Items": {
        "Name": "Specialty_Group",
        "Label": "Specialty Group",
        "Description": "The primary overarching medical field or expertise area to which the healthcare provider belongs. Uses the list of global specialties.",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "CHAR(2)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(40)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
        ],
        "Values": [],
    },
    "Medical Degree Items": {
        "Name": "Medical_Degree",
        "Label": "Medical Degree",
        "Description": "The primary medical qualification or degree obtained.",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "VARCHAR(4)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(40)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
        ],
        "Values": [],
    },
    "HCP Status Items": {
        "Name": "HCP_Status",
        "Label": "HCP Status",
        "Description": "Indicates whether the healthcare professional is currently active and working or not.",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "CHAR(4)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(20)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
        ],
        "Values": [],
    },
    "Level Items": {
        "Name": "Level",
        "Label": "Level",
        "Description": "Indicates the level of importance of this individual to the company, where level 5 indicates the highest level of importance. Can be used to drive business rules. For example: You may want to limit personalized promotions to levels 3 and below. You may also require a single relationship owner for level 5.",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "SMALLINT",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(20)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
        ],
        "Values": [],
    },
    "Adopter Type Items": {
        "Name": "Adopter_Type",
        "Label": "Adopter Type",
        "Description": "A categorization of the individual based on their willingness and speed to adopt new medical technologies, treatments, practices, or products.",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "CHAR(4)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(20)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
        ],
        "Values": [],
    },
    "Address Status Items": {
        "Name": "Address_Status",
        "Label": "Address Status",
        "Description": "Indicates whether this address is currently usable for contact purposes.",
        "Attributes": [
            {
                "Name": "name",
                "Label": "Name",
                "Data Type": "VARCHAR(4)",
                "Description": None,
            },
            {
                "Name": "label",
                "Label": "Label",
                "Data Type": "VARCHAR(20)",
                "Description": None,
            },
            {
                "Name": "description",
                "Label": "Description",
                "Data Type": "VARCHAR",
                "Description": None,
            },
        ],
        "Values": [],
    },
}


DATA_TYPE_MAPPING_POSTGRES: dict[str, str] = {
    "Text 10": "VARCHAR(10)",
    "Text 15": "VARCHAR(15)",
    "Text 20": "VARCHAR(20)",
    "Text 25": "VARCHAR(25)",
    "Text 40": "VARCHAR(40)",
    "Text 50": "VARCHAR(50)",
    "Text 80": "VARCHAR(80)",
    "Text 100": "VARCHAR(100)",
    "Boolean": "BOOLEAN",
    "hcp.language.Picklist": "CHAR(2)",
    "hcp.country.Picklist": "CHAR(2)",
    "hcp.state.Picklist": "VARCHAR(6)",
    "hcp.hcp_type.Picklist": "CHAR(4)",
    "hcp.spec_1.Picklist": "VARCHAR(4)",
    "hcp.all_spec.Multivalue Picklist": "VARCHAR(4)[]",
    "hcp.spec_group_1.Picklist": "CHAR(2)",
    "hcp.all_spec_group.Multivalue Picklist": "CHAR(2)[]",
    "hcp.degree_1.Picklist": "VARCHAR(4)",
    "hcp.all_degree.Multivalue Picklist": "VARCHAR(4)[]",
    "hcp.status.Picklist": "VARCHAR(4)",
    "hcp.level.Picklist": "SMALLINT",
    "hcp.adopter_type.Picklist": "CHAR(4)",
    "address.country.Picklist": "CHAR(2)",
    "address.state.Picklist": "VARCHAR(6)",
    "address.status.Picklist": "VARCHAR(4)",
    "Entity (HCP)": "INT",
    "SMALLINT": "SMALLINT",
    "CHAR(2)": "CHAR(2)",
    "CHAR(3)": "CHAR(3)",
    "CHAR(4)": "CHAR(4)",
    "VARCHAR(4)": "VARCHAR(4)",
    "VARCHAR(6)": "VARCHAR(6)",
    "VARCHAR(20)": "VARCHAR(20)",
    "VARCHAR(40)": "VARCHAR(40)",
    "VARCHAR(80)": "VARCHAR(80)",
    "VARCHAR": "VARCHAR",
}

type Kernel = dict[str, dict[str, dict[str, str | list[dict[str, str]]]]]


def run() -> None:
    input_folder = Path().cwd() / "input"
    output_folder = Path().cwd() / "output"

    for file in input_folder.glob("*.xlsx"):
        wb = load_workbook(filename=file, data_only=True)

        kernel = {}
        kernel = _add_picklists(wb, kernel)
        kernel = _add_entities(wb, kernel)
        kernel = _add_attributes(wb, kernel)
        kernel = _add_license(wb, kernel)

        output_subfolder = output_folder / file.parts[-1].replace(".xlsx", "").replace(
            ".", "_"
        ).replace("-", "_")
        output_subfolder.mkdir(exist_ok=True)

        _create_output(output_subfolder, kernel)


def _add_picklists(wb: Workbook, kernel: Kernel) -> Kernel:
    """improve me"""
    picklists = set(wb.get_sheet_names()) - {  # noqa
        SHEET_NAME_ATTRIBUTES,
        SHEET_NAME_ENTITIES,
        SHEET_NAME_LICENSE,
    }

    kernel[f"Picklist {SHEET_NAME_ENTITIES}"] = {}

    for picklist in picklists:
        definition = PICKLIST_ENTITIES_AND_ATTRIBUTES[picklist]
        name = definition.pop("Name")
        kernel[f"Picklist {SHEET_NAME_ENTITIES}"][name] = definition

        _sheet_iterator(wb, picklist, kernel, _add_picklist_values, name)

    return kernel


def _add_entities(wb: Workbook, kernel: Kernel) -> Kernel:
    kernel[SHEET_NAME_ENTITIES] = {}
    return _sheet_iterator(wb, SHEET_NAME_ENTITIES, kernel, _add_entity)


def _add_attributes(wb: Workbook, kernel: Kernel) -> Kernel:
    for entity in kernel[SHEET_NAME_ENTITIES].values():
        entity[SHEET_NAME_ATTRIBUTES] = []
    return _sheet_iterator(wb, SHEET_NAME_ATTRIBUTES, kernel, _add_attribute)


def _add_license(wb: Workbook, kernel: Kernel) -> Kernel:
    _, version, *date = _parse_cell_value(
        wb[SHEET_NAME_LICENSE].cell(row=1, column=1)
    ).split(",")
    date = [part.strip() for part in date]
    kernel[SHEET_NAME_LICENSE] = {
        "Version": version.strip(),
        "Date": ",".join(date),
        "Title": _parse_cell_value(wb[SHEET_NAME_LICENSE].cell(row=2, column=1)),
        "Text": _parse_cell_value(wb[SHEET_NAME_LICENSE].cell(row=3, column=1)),
    }
    return kernel


def _create_output(output_folder: Path, kernel: Kernel) -> None:
    with open(output_folder / "yaml.yaml", "w") as yaml_file:
        yaml.dump(dict(kernel), yaml_file, allow_unicode=True)

    postgresql_query = _build_postgresql_query(kernel)
    with open(output_folder / "postgres.sql", "w") as postgres_file:
        postgres_file.write(postgresql_query)


def _sheet_iterator(
    wb: Workbook, sheet_name: str, kernel: Kernel, add_row: Callable, name: str = None
) -> Kernel:
    idx = -1
    fields = None
    for i, row in enumerate(wb[sheet_name].rows):
        if sum(cell.value is not None for cell in row) == 1:
            idx = i
            continue

        if i == idx + 1:
            fields = [item.value for item in row if item.value is not None]
            continue

        if fields is None:
            raise RuntimeError("Did not find a row that defines the fields.")

        if all(cell.value is None for cell in row):
            break

        kernel = add_row(row, kernel, fields, name or sheet_name)

    return kernel


def _add_picklist_values(
    row: tuple[Cell], kernel: Kernel, fields: list[str], sheet_name: str
) -> Kernel:
    kernel[f"Picklist {SHEET_NAME_ENTITIES}"][sheet_name]["Values"].append(
        {field.lower(): _parse_cell_value(cell) for field, cell in zip(fields, row)}
    )
    return kernel


def _add_entity(row: tuple[Cell], kernel: Kernel, fields: list[str], *args) -> Kernel:
    kernel[SHEET_NAME_ENTITIES][_parse_cell_value(row[0])] = {
        field: _parse_cell_value(cell)
        for field, cell in zip(fields[1:], row[1 : len(fields)])
    }

    return kernel


def _add_attribute(
    row: tuple[Cell], kernel: Kernel, fields: list[str], *args
) -> Kernel:
    if row[0].value in kernel[SHEET_NAME_ENTITIES].keys():
        kernel[SHEET_NAME_ENTITIES][row[0].value][SHEET_NAME_ATTRIBUTES].append(
            {
                field: _parse_cell_value(cell)
                for field, cell in zip(fields[1:], row[1 : len(fields)])
            }
        )

    return kernel


def _parse_cell_value(cell: Cell) -> str | None:
    if isinstance(cell.value, str):
        return None if cell.value == "N/A" else cell.value.strip()
    elif cell.value is None:
        return None
    elif isinstance(cell.value, int):
        return str(cell.value)
    else:
        raise ValueError(f"Found unexpected type {type(cell.value)} in {cell}")


def _build_postgresql_query(kernel: Kernel) -> str:
    query = ""
    for name, definition in kernel[SHEET_NAME_ENTITIES].items():
        name = _escape_sql_keyword(name.lower())
        query += f"""DROP TABLE IF EXISTS {name}
;
CREATE TABLE {name} (
    {_build_sql_columns(definition[SHEET_NAME_ATTRIBUTES], name)}
)
;
COMMENT ON TABLE {name} IS '{_build_sql_table_comment(definition)}'
;
{_build_sql_column_comments(definition[SHEET_NAME_ATTRIBUTES], name)}
;

"""

    for name, definition in kernel["Picklist Entities"].items():
        name = f"picklist_{name.lower()}"
        query += f"""DROP TABLE IF EXISTS {name}
;
CREATE TABLE {name} (
    {_build_sql_columns(definition[SHEET_NAME_ATTRIBUTES], name)}
)
;
{_build_sql_column_comments(definition[SHEET_NAME_ATTRIBUTES], name)}
;
{_insert_values(definition, name)}
;

"""

    return query


def _build_sql_columns(attributes: list[dict], table_name: str) -> str:
    return ",\n    ".join(
        [
            f"{_escape_sql_keyword(attribute['Name'].lower())} {
                DATA_TYPE_MAPPING_POSTGRES[
                    table_name + '.' + attribute['Name'].lower() + '.' + attribute['Data Type']
                ]
                if attribute['Data Type'] in ("Picklist", "Multivalue Picklist") else
                DATA_TYPE_MAPPING_POSTGRES[attribute['Data Type']]
            }"
            for attribute in attributes
        ]
    )


def _build_sql_column_comments(attributes: list[dict], table_name: str) -> str:
    return "\n;\n".join(
        [
            f"COMMENT ON COLUMN {table_name}.{attribute['Name'].lower()} IS"
            f" '{_build_sql_comment(attribute, exclude=('Name', 'Data Type'))}'"
            for attribute in attributes
        ]
    )


def _build_sql_table_comment(definition: dict) -> str:
    return _build_sql_comment(definition, (SHEET_NAME_ATTRIBUTES,))


def _build_sql_comment(definition: dict, exclude: tuple = tuple()) -> str:
    return ", ".join(
        [
            f"{key}: {value.replace("'", "''")}"
            for key, value in definition.items()
            if key not in exclude and value is not None
        ]
    )


def _insert_values(definition: dict, table_name: str) -> str:
    fields = sorted(
        [attribute["Name"] for attribute in definition[SHEET_NAME_ATTRIBUTES]]
    )
    query = f"INSERT INTO {table_name} ({", ".join(fields)}) VALUES"
    values = ",\n    ".join(_parse_row(row) for row in definition["Values"])
    return f"{query}\n    {values}"


def _parse_row(row: dict[str, str | int | None]) -> str:
    def _parse_value(value):
        if isinstance(value, str):
            return f"'{value.replace("'", "''")}'"
        if isinstance(value, int):
            return str(value)
        if value is None:
            return "null"
        raise ValueError(f"Found unexpected type {type(value)} for value {value}")

    sorted_row = dict(sorted(row.items()))
    return f"({", ".join([_parse_value(value) for value in sorted_row.values()])})"


def _escape_sql_keyword(string: str) -> str:
    sql_keywords = _get_sql_keywords()

    if string.upper() in sql_keywords:
        string = f'"{string}"'

    return string


@lru_cache
def _get_sql_keywords() -> tuple[str, ...]:
    with open(Path.cwd() / "resources" / "sql_keywords_postgresql.yaml") as stream:
        return tuple(yaml.safe_load(stream)["sql_keywords"])
